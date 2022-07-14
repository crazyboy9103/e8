import numpy as np
import json
from tqdm import tqdm
import time, datetime

def getTimestamp():
    timezone=60*60*9
    utc_timestamp = int(time.time()+timezone)
    date = datetime.datetime.fromtimestamp(utc_timestamp).strftime("%Y-%m-%d %H:%M:%S")
    return date

def compute_iou(cand_box, gt_box):
    # Calculate intersection areas
    x1 = np.maximum(cand_box[0], gt_box[0])
    y1 = np.maximum(cand_box[1], gt_box[1])
    x2 = np.minimum(cand_box[2], gt_box[2])
    y2 = np.minimum(cand_box[3], gt_box[3])

    intersection = np.maximum(x2 - x1, 0) * np.maximum(y2 - y1, 0)

    cand_box_area = (cand_box[2] - cand_box[0]) * (cand_box[3] - cand_box[1])
    gt_box_area = (gt_box[2] - gt_box[0]) * (gt_box[3] - gt_box[1])
    union = cand_box_area + gt_box_area - intersection

    iou = intersection / union
    return iou



def analysis(metrics):
    logs = {}
    for image_name, result in tqdm(metrics.items(), desc="reading image result"):
        if isinstance(result, int) or isinstance(result, str):
            continue
        if image_name not in logs:
            logs[image_name] = {}

        for class_name, stats in result.items():
            if class_name not in logs[image_name]:
                logs[image_name][class_name] = {}
                logs[image_name][class_name]["gt_label"] = []
                logs[image_name][class_name]["label"] = []
                logs[image_name][class_name]["gt_bbox"] = []
                logs[image_name][class_name]["bbox"] = []
                logs[image_name][class_name]["conf"] = []
                logs[image_name][class_name]["iou"] = []
                logs[image_name][class_name]["correct"] = []
                logs[image_name][class_name]["time"] = []
                logs[image_name][class_name]["FP"] = []
                logs[image_name][class_name]["FN"] = []
                logs[image_name][class_name]["TP"] = []


            gt_bbox = stats['gt_bbox']
            gt_label = stats['gt_label']
            pred_bbox = stats['bbox']
            pred_label = stats['label']
            conf = stats['conf']

            #pred_bbox_tensor = torch.tensor(pred_bbox)
            #conf_tensor = torch.tensor(conf)

            #keep_idx = torchvision.ops.nms(pred_bbox_tensor, conf_tensor, 0.1).tolist()
            #if len(keep_idx) > 3:
            #    keep_idx = keep_idx[:3]
            
            #pred_bbox = [pred_bbox[idx] for idx in keep_idx]
            #pred_label = [pred_label[idx] for idx in keep_idx]
            #conf = [conf[idx] for idx in keep_idx]

            
            for i in range(len(gt_bbox)):
                best_iou = -9999
                best_idx = None
                for j in range(len(pred_bbox)):
                    iou = compute_iou(gt_bbox[i], pred_bbox[j])
                    if iou > best_iou:
                        best_iou = iou
                        best_idx = j
                #print(best_idx)
                #assert best_idx
                logs[image_name][class_name]["gt_label"].append(gt_label[i])
                logs[image_name][class_name]["gt_bbox"].append(gt_bbox[i])
                logs[image_name][class_name]["label"].append(pred_label[best_idx])
                logs[image_name][class_name]["bbox"].append(pred_bbox[best_idx])
                logs[image_name][class_name]["conf"].append(conf[best_idx])
                logs[image_name][class_name]["iou"].append(best_iou)

                correct = pred_label[best_idx] == gt_label[i] if best_iou > 0.5 else False
                logs[image_name][class_name]["correct"].append(correct)
                logs[image_name][class_name]["time"].append(getTimestamp())
    
                FP = (pred_label[best_idx] == gt_label[i]) == True and best_iou < 0.5
                FN = (pred_label[best_idx] == gt_label[i]) == False
                TP = (pred_label[best_idx] == gt_label[i]) == True and best_iou > 0.5
                logs[image_name][class_name]["FP"].append(int(FP))
                logs[image_name][class_name]["FN"].append(int(FN))
                logs[image_name][class_name]["TP"].append(int(TP))
    return logs

from itertools import accumulate
from scipy import integrate
from openpyxl import Workbook
def write_to_excel(metrics):
    labels = json.load(open("labels.json", "r"))
    labels = list(labels.keys())
    wb = Workbook(write_only=True)

    ws = wb.create_sheet("SSD")
    # worksheets = {}
    
    header = ["image_name", "time", "correct", "gt_label", "gt_bbox", "label", "bbox", "conf", "iou", "cum_TP", "cum_FN", "cum_FP", "recall", "precision", "average_precision", "AP", "avg_iou"]
    # for label in labels:
    #     ws = wb.create_sheet(label)
    #     worksheets[label] = ws
    #     ws.append(header)

    analysis_result = {label: [] for label in labels}
    logs = analysis(metrics)
    for image_name, result in tqdm(logs.items(), desc="image analysis"):
        if isinstance(result, int) or isinstance(result, str):
            continue

        for class_name, stats in result.items():
            for i in range(len(stats['label'])):
                analysis_result[class_name].append([image_name, stats["time"][i], stats["correct"][i], stats["gt_label"][i], list(map(int, stats["gt_bbox"][i])), stats["label"][i], list(map(int, stats["bbox"][i])), round(stats["conf"][i], 5), round(stats["iou"][i],2), stats["FP"][i], stats["FN"][i], stats["TP"][i]])

    del logs

    APs = {}
    mious = {}
    
    epsilon = 1e-6
    for label, result in tqdm(analysis_result.items(), desc="writing to excel"):
        # ws = worksheets[label]
        ws.append([label])
        ws.append(header)
        result = sorted(result, key=lambda x: x[7], reverse=True)
        ious = [item[-4] for item in result]
        mean_iou = sum(ious)/len(ious)
        TP = [item[-1] for item in result]
        FN = [item[-2] for item in result]
        FP = [item[-3] for item in result]
    
        cum_TP = list(accumulate(TP))
        cum_FN = list(accumulate(FN))
        cum_FP = list(accumulate(FP))
    
        recalls = [tp/(cum_TP[-1] + cum_FN[-1] + epsilon) for tp in cum_TP]
        precisions = [tp/(tp + fp + epsilon) for tp, fp in zip(cum_TP, cum_FP)]
        average_precisions = integrate.cumtrapz([1]+precisions, [0]+recalls)
        recalls = list(map(lambda x: round(x, 4), recalls))
        precisions = list(map(lambda x: round(x, 4), precisions))
        average_precisions = list(map(lambda x: round(x, 4), average_precisions))
        for i, (res, cum_tp, cum_fn, cum_fp, rec, prec, avg_prec) in tqdm(enumerate(zip(result, cum_TP, cum_FN, cum_FP, recalls, precisions, average_precisions)), desc=f"{label} calc stats"):
            line = res[:9] + [cum_tp, cum_fn, cum_fp, rec, prec, avg_prec]
            if i == 0:
                APs[label] = average_precisions[-1]
                mious[label] = mean_iou
                line = line + [average_precisions[-1], mean_iou]
            line = list(map(str, line))
            ws.append(line)
    
    # ws = wb.create_sheet("Stats")
    ws.append(["Stats"])
    ws.append(["Class", "AP", "mIoU", "Final_mAP", "Final_mIoU"])
    for i, label in tqdm(enumerate(labels), desc="total stats"):
        AP = APs[label]
        miou = mious[label]
        line = [label, AP, miou]
        if i == 0:
            line = line + [sum(APs.values())/len(APs), sum(mious.values())/len(mious)]
        line = list(map(str, line))
        ws.append(line) 
    print("Saving test.xlsx")
    wb.save("test.xlsx")
    wb.close()


if __name__ == "__main__":
    print("Reading detailed_metrics.json...")
    f = open("detailed_metrics.json", "r")
    metrics = json.load(f)
    f.close()
    print("Finished reading")
    write_to_excel(metrics)

