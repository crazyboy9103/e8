from re import template
import torchvision.transforms as transforms
from torchvision.datasets import ImageFolder
import torch
from torch.utils.data import DataLoader
import torchvision.models as models
import torchvision
import numpy as np
import matplotlib.pyplot as plt
import torch.nn as nn
import argparse
import json
from tqdm import tqdm
import os
import csv

class ImageFolderWithPaths(ImageFolder):
    """Custom dataset that includes image file paths. Extends
    torchvision.datasets.ImageFolder
    """

    # override the __getitem__ method. this is the method that dataloader calls
    def __getitem__(self, index):
        # this is what ImageFolder normally returns 
        original_tuple = super(ImageFolderWithPaths, self).__getitem__(index)
        # the image file path
        path = self.imgs[index][0]
        # make a new tuple that includes original and the path
        tuple_with_path = (original_tuple + (path,))
        return tuple_with_path


def getTimestamp():
    import time, datetime
    timezone = 60*60*9 # seconds * minutes * utc + 9
    utc_timestamp = int(time.time() + timezone)
    date = datetime.datetime.fromtimestamp(utc_timestamp).strftime('%Y-%m-%d %H:%M:%S')
    return date

parser = argparse.ArgumentParser(description="train efficientnet-b0")
parser.add_argument("--dataset", default="./dataset", type=str, help="dataset path")
parser.add_argument("--model", default="eff_net.pt", type=str, help="model name to load from")
args = parser.parse_args()

transforms = transforms.Compose([
    transforms.ToTensor()
])

test_dir  = args.dataset
dataset = ImageFolderWithPaths(root=test_dir, transform=transforms, target_transform=None)
model_name = args.model.strip(".pt")
idx_filename = f"{model_name}_idx.npy"

if idx_filename not in os.listdir():
    test_idx = np.random.choice(len(dataset), len(dataset)//10, replace=False)
    np.save(idx_filename, test_idx)
else:
    test_idx = np.load(idx_filename)

testset = torch.utils.data.Subset(dataset, test_idx)
f = open(f"test_{model_name}.csv", "w", newline='', encoding="utf-8-sig")
csv_writer = csv.writer(f)
testloader = DataLoader(testset, batch_size=1, shuffle=True, pin_memory=True, num_workers=4)

full_image_names = dataset.imgs
filenames = [full_image_names[idx] for idx in test_idx]
for row in filenames:
    csv_writer.writerow([row[0]])
f.close()

PATH = args.model


def build_net(num_classes):
    net = models.efficientnet_b0(pretrained=True)
    num_ftrs = net.classifier[1].in_features
    net.classifier[1] = nn.Linear(num_ftrs, num_classes)
    return net

net = build_net(3)
net.load_state_dict(torch.load(PATH))
device = torch.device("cuda:0")
net.to(device)
net.eval()

logs = {"start":getTimestamp()}
stats_by_class = {dataset.classes[i]:{"correct":0, "total":0} for i in range(3)} 
labels_by_class = {dataset.classes[i]:[] for i in range(3)}
preds_by_class = {dataset.classes[i]:[] for i in range(3)}

from sklearn.metrics import f1_score, precision_score, recall_score, accuracy_score
with torch.no_grad():
    for img_idx, (image, label, path) in enumerate(tqdm(testloader, desc="evaluating")):
        output = net(image.to(device))

        _, predicted = torch.max(output, 1)
        path = path[0]
        img_name = path
        temp_label = dataset.classes[label.item()]
        temp_predict = dataset.classes[predicted.item()]
        is_correct = temp_label == temp_predict
        stats_by_class[temp_label]["total"] += 1
        stats_by_class[temp_label]["correct"] += int(is_correct)
 
        logs[path] = {"predict":temp_predict, "label": temp_label, "is_correct": is_correct, "time": getTimestamp(),  "class_stats":{}, "final_stats":{}}
        logs[path]["cumul_correct"] = stats_by_class[temp_label]["correct"]
        logs[path]["cumul_total"] = stats_by_class[temp_label]["total"]


        labels_by_class[temp_label].append(temp_label)
        preds_by_class[temp_label].append(temp_predict)


    logs["class_stats"] = {}
    for i in range(3):
        labels = labels_by_class[dataset.classes[i]]
        preds = preds_by_class[dataset.classes[i]]
        logs["class_stats"][dataset.classes[i]] = {"acc": accuracy_score(labels, preds), "f1":f1_score(labels, preds, average="weighted")}

    mean_acc = np.mean(list(logs["class_stats"][dataset.classes[i]]["acc"] for i in range(3)))
    mean_f1 = np.mean(list(logs["class_stats"][dataset.classes[i]]["f1"] for i in range(3)))

    logs["final_stats"] = {"mean_acc":mean_acc, "mean_f1":mean_f1}

logs["end"] = getTimestamp()

json.dump(logs, open("detailed_metrics.json", "w"))
from openpyxl import Workbook
from sklearn.metrics import multilabel_confusion_matrix as mcm
def write_to_excel(logs):
    

    labels = []
    preds = []

    for k, v in labels_by_class.items():
        labels.extend(v)
    for k, v in preds_by_class.items():
        preds.extend(v)


    multi_confusion = mcm(labels, preds, labels=["best", "normal", "faulty"])
    best_conf, norm_conf, fault_conf = multi_confusion


    b_tn, b_fp, b_fn, b_tp = best_conf.ravel()
    n_tn, n_fp, n_fn, n_tp = norm_conf.ravel()
    f_tn, f_fp, f_fn, f_tp = fault_conf.ravel()

    b_prec, b_recall = b_tp/(b_tp+b_fp), b_tp/(b_tp+b_fn)
    b_f1 = 2 * (b_prec * b_recall)/(b_prec + b_recall)
    b_acc = (b_tp+b_tn) / (b_tp+b_tn+b_fp+b_fn)

    n_prec, n_recall = n_tp/(n_tp+n_fp), n_tp/(n_tp+n_fn)
    n_f1 = 2 * (n_prec * n_recall)/(n_prec + n_recall)
    n_acc = (n_tp+n_tn) / (n_tp+n_tn+n_fp+n_fn)

    f_prec, f_recall = f_tp/(f_tp+f_fp), f_tp/(f_tp+f_fn)
    f_f1 = 2 * (f_prec * f_recall)/(f_prec + f_recall)
    f_acc = (f_tp+f_tn) / (f_tp+f_tn+f_fp+f_fn)

    wb = Workbook()
    ws = wb.active

    model_name = args.model.strip(".pt")

    ws.append([model_name+"best_tn", model_name+"best_fp", model_name+"best_fn", model_name+"best_tp", model_name+"best_acc", model_name+"best_f1"])
    ws.append([b_tn, b_fp, b_fn, b_tp, b_acc, b_f1])
    ws.append([model_name+"normal_tn", model_name+"normal_fp", model_name+"normal_fn", model_name+"normal_tp", model_name+"normal_acc", model_name+"normal_f1"])
    ws.append([n_tn, n_fp, n_fn, n_tp, n_acc, n_f1])
    ws.append([model_name+"faulty_tn", model_name+"faulty_fp", model_name+"faulty_fn", model_name+"faulty_tp", model_name+"faulty_acc", model_name+"faulty_f1"])
    ws.append([f_tn, f_fp, f_fn, f_tp, f_acc, f_f1])
    ws.append(["image_name", "predict", "label", "cumul_correct", "cumul_total", "", "Class", "Accuracy", "F1", "", "Final mean Acc", "mean F1"])

    analysis_result = []
    for key, value in logs.items():
        if key not in ["final_stats", "end", "start", "class_stats"]:
            img_name = key

            try:
                analysis_result.append([img_name,str(value["predict"]),str(value["label"]),str(value["cumul_correct"]),str(value["cumul_total"])])

            except:
                continue

    model_name = args.model.strip(".pt")

    print(f"Eval started : {logs['start']}, Eval ended : {logs['end']}")
    for i in range(3):
        print(f"Class {dataset.classes[i]}, Accuracy: {logs['class_stats'][dataset.classes[i]]['acc']}, F1: {logs['class_stats'][dataset.classes[i]]['f1']}")
        analysis_result[i].extend(["", model_name+dataset.classes[i], logs['class_stats'][dataset.classes[i]]['acc'], logs['class_stats'][dataset.classes[i]]['f1']])

    print(f"Final mean Acc : {logs['final_stats']['mean_acc']}, mean F1 : {logs['final_stats']['mean_f1']}")
    analysis_result[0].extend(["", logs['final_stats']['mean_acc'], logs['final_stats']['mean_f1']])

    for row in analysis_result:
        ws.append(row)
    wb.save(model_name+"_test.xlsx")
    print(model_name+"_test.xlsx saved")

write_to_excel(logs)

import openpyxl
file_list = ["eff_net_eff_net_window_test", "eff_net_eff_net_rebar_test", "eff_net_eff_net_peel_test", "eff_net_eff_net_living_test", "eff_net_eff_net_ground_test", "eff_net_eff_net_finish_test", "eff_net_eff_net_crack_test"]
path_list = [i+".xlsx" for i in file_list]
best_data = []
log_data = []
class_data = [['model', 'Class', 'Accuracy', 'F1']]
f1_data = [] # 평균 계산을 위한 컬럼명 제외

for path in path_list:
    model = path.split("/")[-1].split(".")[0].replace("eff_net_eff_net_", "")
    wb = openpyxl.load_workbook(path)
    ws = wb['Sheet']
    result = []
    for row in ws.iter_rows(ws.min_row, ws.max_row, values_only =True):
        result.append(list(row))

    for idx, row in enumerate(result[0:6]):
        if idx == 0:
            column_name = ["model"]
            column_name.extend(row[0:6])
            best_data.append(column_name)
        else:
            data = [model]
            data.extend(row[0:6])
            best_data.append(data)

    for idx, row in enumerate(result[6:]):
        log_data.append(row[:5])
        if idx in [1]:
            f1_data.append(row[10:])
        if idx in [1, 2 ,3]:
            data = [model]
            data.extend(row[6:9])
            class_data.append(data)

acc = 0
f1 = 0
for row in f1_data:
    acc += row[0]
    f1 += row[1]

for idx, row in enumerate(best_data):
    data = [""]
    data.extend(row)
    log_data[idx].extend(data)

for idx, row in enumerate(class_data):
    data = [""]
    data.extend(row)
    log_data[idx].extend(data)

f1_average = [["", "Final mean Acc", "mean F1"],["", acc/len(file_list), f1/len(file_list)]]
for idx, row in enumerate(f1_average):
    log_data[idx].extend(row)

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
for row in log_data:
    ws.append(row)
wb.save("eff_net_result_test.xlsx")
