import numpy as np
import json
from tqdm import tqdm
from sklearn.metrics import average_precision_score
from openpyxl import Workbook
from itertools import accumulate
from scipy import integrate

# Load logs
print("Reading detailed_metrics.json...")
f = open("detailed_metrics.json", "r")
logs = json.load(f)
f.close()
print("Finished reading")

# Labels
labels = json.load(open("labels.json", "r"))
labels = list(labels.keys())
# Open workbook
wb = Workbook(write_only=True)
ws = wb.create_sheet("MRCNN")
# worksheets = {}
header = ["image_name", "time", "correct", "gt_label", "label", "conf", "iou", "cum_TP", "cum_FN", "cum_FP", "recall", "precision", "average_precision", "AP", "avg_iou"]
# for label in labels:
#     ws = wb.create_sheet(label)
#     worksheets[label] = ws
#     ws.append(header)


analysis_result = {label: [] for label in labels}
for image_name, result in tqdm(logs.items(), desc=f"image analysis"):
    if isinstance(result, int) or isinstance(result, str):
        continue

    for class_name, stats in result.items():
        for i in range(len(stats['label'])):
            analysis_result[class_name].append([image_name, stats["time"][i], stats["correct"][i], stats["gt_label"][i], stats["label"][i], stats["conf"][i], stats["iou"][i], stats["FP"][i], stats["FN"][i], stats["TP"][i]])

APs = {}
mious = {}   

epsilon = 1e-6
for label, result in tqdm(analysis_result.items(), desc="writing to excel"):
    # ws = worksheets[label]
    ws.append([label])
    ws.append(header)
    result = sorted(result, key=lambda x: x[5], reverse=True)
    ious = [item[-4] for item in result]
    mean_iou = sum(ious)/len(ious)
    TP = [item[-1] for item in result]
    FN = [item[-2] for item in result]
    FP = [item[-3] for item in result]

    cum_TP = list(accumulate(TP))
    cum_FN = list(accumulate(FN))
    cum_FP = list(accumulate(FP))

    # Compute cum recalls precisions
    recalls = [tp/(cum_TP[-1] + cum_FN[-1] + epsilon) for tp in cum_TP]
    precisions = [tp/(tp + fp + epsilon) for tp, fp in zip(cum_TP, cum_FP)]
    average_precisions = integrate.cumtrapz([1]+precisions, [0]+recalls)

    # Round 4th decimal point to avoid large string byte size
    recalls = list(map(lambda x: round(x, 4), recalls))
    precisions = list(map(lambda x: round(x, 4), precisions))
    average_precisions = list(map(lambda x: round(x, 4), average_precisions))

    for i, (res, cum_tp, cum_fn, cum_fp, rec, prec, avg_prec) in tqdm(enumerate(zip(result, cum_TP, cum_FN, cum_FP, recalls, precisions, average_precisions)),  desc=f"{label} calc stats"):
        line = res[:7] + [cum_tp, cum_fn, cum_fp, rec, prec, avg_prec]
        if i == 0:
            APs[label] = average_precisions[-1]
            mious[label] = mean_iou
            line = line + [average_precisions[-1], mean_iou]
        line = list(map(str, line))
        ws.append(line)

# ws = wb.create_sheet("Stats")
ws.append(["Stats"])
ws.append(["Class", "AP", "mIoU", "Final_mAP", "Final_mIoU"])
for i, label in enumerate(labels):
    AP = APs[label]
    miou = mious[label]
    line = [label, AP, miou]
    if i == 0:
        line = line + [sum(APs.values())/len(APs), sum(mious.values())/len(mious)]
    line = list(map(str, line))
    ws.append(line) 

print("Saving mrcnn_test.xlsx")
wb.save("mrcnn_test.xlsx")
wb.close()