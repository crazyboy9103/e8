import numpy as np
import json
from tqdm import tqdm
import time, datetime
def getTimestamp():
    timezone=60*60*9
    utc_timestamp = int(time.time()+timezone)
    date = datetime.datetime.fromtimestamp(utc_timestamp).strftime("%Y-%m-%d %H:%M:%S")
    return date

def compute_iou(cand_box, gt_box):
    # Calculate intersection areas
    x1 = np.maximum(cand_box[0], gt_box[0])
    y1 = np.maximum(cand_box[1], gt_box[1])
    x2 = np.minimum(cand_box[2], gt_box[2])
    y2 = np.minimum(cand_box[3], gt_box[3])

    intersection = np.maximum(x2 - x1, 0) * np.maximum(y2 - y1, 0)

    cand_box_area = (cand_box[2] - cand_box[0]) * (cand_box[3] - cand_box[1])
    gt_box_area = (gt_box[2] - gt_box[0]) * (gt_box[3] - gt_box[1])
    union = cand_box_area + gt_box_area - intersection

    iou = intersection / union
    return iou

print("Reading detailed_metrics.json...")
f = open("detailed_metrics.json", "r")
metrics = json.load(f)
f.close()
print("Finished reading")

def analysis(metrics):
    logs = {}
    print(f"Eval started : {metrics['start']}, Eval ended : {metrics['end']}")
    for image_name, result in tqdm(metrics.items(), desc="reading image result"):
        if isinstance(result, int) or isinstance(result, str):
            continue
        if image_name not in logs:
            logs[image_name] = {}

        for class_name, stats in result.items():
            if class_name not in logs[image_name]:
                logs[image_name][class_name] = {}
                logs[image_name][class_name]["gt_label"] = []
                logs[image_name][class_name]["label"] = []
                logs[image_name][class_name]["gt_bbox"] = []
                logs[image_name][class_name]["bbox"] = []
                logs[image_name][class_name]["conf"] = []
                logs[image_name][class_name]["iou"] = []
                logs[image_name][class_name]["correct"] = []
                logs[image_name][class_name]["time"] = []
                logs[image_name][class_name]["FP"] = []
                logs[image_name][class_name]["FN"] = []
                logs[image_name][class_name]["TP"] = []
        #logs[image_name][class_name]["TN"] = []

            gt_bbox = stats['gt_bbox']
            gt_label = stats['gt_label']
            pred_bbox = stats['bbox']
            pred_label = stats['label']
            conf = stats['conf']
            for i in range(len(gt_bbox)):
                for j in range(len(pred_bbox)):
                    iou = compute_iou(gt_bbox[i], pred_bbox[j])
                    logs[image_name][class_name]["gt_label"].append(gt_label[i])
                    logs[image_name][class_name]["gt_bbox"].append(gt_bbox[i])
                    logs[image_name][class_name]["label"].append(pred_label[j])
                    logs[image_name][class_name]["bbox"].append(pred_bbox[j])
                    logs[image_name][class_name]["conf"].append(conf[j])
                    logs[image_name][class_name]["iou"].append(iou)

                    correct = pred_label[j] == gt_label[i] if iou > 0.5 else False
                    logs[image_name][class_name]["correct"].append(correct)
                    logs[image_name][class_name]["time"].append(getTimestamp())
        
                    FP = (pred_label[j] == gt_label[i]) == True and iou < 0.5
                    FN = (pred_label[j] == gt_label[i]) == False
                    TP = (pred_label[j] == gt_label[i]) == True and iou > 0.5
                    logs[image_name][class_name]["FP"].append(int(FP))
                    logs[image_name][class_name]["FN"].append(int(FN))
                    logs[image_name][class_name]["TP"].append(int(TP))
    return logs

from itertools import accumulate
from scipy import integrate
def write_to_excel(metrics):
    from openpyxl import Workbook
    correct_by_class = {}
    conf_by_class = {}
    iou_by_class = {}
    
    labels = json.load(open("labels.json", "r"))
    labels = list(labels.keys())
    wb = Workbook()
    assert "Sheet" in wb.sheetnames
    worksheets = {}
    
    header = ["image_name", "time", "correct", "gt_label", "gt_bbox", "label", "bbox", "conf", "iou", "cum_TP", "cum_FN", "cum_FP", "recall", "precision", "average_precision", "AP", "avg_iou"]
    for label in labels:
        ws = wb.create_sheet(label)
        worksheets[label] = ws
        ws.append(header)

    analysis_result = {label: [] for label in labels}
    logs = analysis(metrics)
    for image_name, result in tqdm(logs.items(), desc="image analysis"):
        if isinstance(result, int) or isinstance(result, str):
            continue

        for class_name, stats in result.items():
            for i in range(len(stats['label'])):
                analysis_result[class_name].append([image_name, stats["time"][i], stats["correct"][i], stats["gt_label"][i], stats["gt_bbox"][i], stats["label"][i], stats["bbox"][i], stats["conf"][i], stats["iou"][i], stats["FP"][i], stats["FN"][i], stats["TP"][i]])

    APs = {}
    mious = {}
    
    epsilon = 1e-6
    for label, result in tqdm(analysis_result.items(), desc="writing to excel"):
        ws = worksheets[label]
        result = sorted(result, key=lambda x: x[7], reverse=True)
        ious = [item[-4] for item in result]
        mean_iou = sum(ious)/len(ious)
        TP = [item[-1] for item in result]
        FN = [item[-2] for item in result]
        FP = [item[-3] for item in result]
    
        cum_TP = list(accumulate(TP))
        cum_FN = list(accumulate(FN))
        cum_FP = list(accumulate(FP))
    
        recalls = [tp/(cum_TP[-1] + cum_FN[-1] + epsilon) for tp in cum_TP]
        precisions = [tp/(cum_TP[-1] + cum_FP[-1] + epsilon) for tp in cum_TP]
        average_precisions = integrate.cumtrapz([1]+precisions, [0]+recalls)
    
        for i, (res, cum_tp, cum_fn, cum_fp, rec, prec, avg_prec) in enumerate(zip(result, cum_TP, cum_FN, cum_FP, recalls, precisions, average_precisions)):
            line = res[:9] + [cum_tp, cum_fn, cum_fp, rec, prec, avg_prec]
            if i == 0:
                APs[label] = average_precisions[-1]
                mious[label] = mean_iou
                line = line + [average_precisions[-1], mean_iou]
            line = list(map(str, line))
            ws.append(line)
    
    ws = wb["Sheet"]
    ws.title = "Stats"
    ws.append(["Class", "AP", "average IoU", "mAP", "mIoU"])
    for i, label in enumerate(labels):
        AP = APs[label]
        miou = mious[label]
        line = [label, AP, miou]
        if i == 0:
            line = line + [sum(APs.values())/len(APs), sum(mious.values())/len(mious)]
        line = list(map(str, line))
        ws.append(line) 
    wb.save("test.xlsx")
    wb.close()

write_to_excel(metrics)

